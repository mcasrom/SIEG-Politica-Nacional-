#!/usr/bin/env python3
"""
export_semanal.py
SIEG – Centro OSINT · Política Nacional

Genera exportación semanal en Excel con múltiples hojas:
- Noticias (últimos 7 días)
- Tendencias diarias
- Narrativas por partido
- Valoración de líderes
- Resumen ejecutivo

Envía el archivo por Telegram y lo deja en data/exports/

Cron: domingos 06:00
0 6 * * 0 cd ~/SIEG-Politica-Nacional && source venv/bin/activate && python3 scripts/export_semanal.py >> logs/export.log 2>&1

Autor : M. Castillo · mybloggingnotes@gmail.com
© 2026 M. Castillo – Todos los derechos reservados
"""

import os
import json
import sqlite3
import requests
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.expanduser("~/SIEG-Politica-Nacional")
DB_PATH    = os.path.join(BASE_DIR, "data", "processed", "noticias.db")
CFG_PATH   = os.path.join(BASE_DIR, "config", "politica_config.json")
LOG_PATH   = os.path.join(BASE_DIR, "logs", "pipeline.log")
OUT_DIR    = os.path.join(BASE_DIR, "data", "exports")
BOT_TOKEN  = "8789958560:AAGRB5opW11gL6m13cXwUAjJcr_bZN2Y9fM"
CHANNEL    = "@sieg_politica"

# Colores corporativos SIEG OSINT
COLOR_HEADER  = "0A0E0A"  # negro OSINT
COLOR_GREEN   = "00CC33"  # verde terminal
COLOR_ACCENT  = "1A3A1A"  # verde oscuro
COLOR_POS     = "2CA02C"
COLOR_NEG     = "D62728"
COLOR_NEU     = "888888"

def log(msg):
    ts = datetime.now().isoformat()
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] [EXPORT] {msg}\n")
    print(f"[EXPORT] {msg}")

def estilo_header(ws, row, cols, color_bg=COLOR_HEADER, color_fg="00FF41"):
    """Aplica estilo de cabecera a una fila."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=color_fg, name="Courier New")
        cell.fill = PatternFill("solid", fgColor=color_bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            bottom=Side(style="thin", color=COLOR_GREEN)
        )

def autofit(ws):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

def crear_excel(df_noticias, df_tendencias, cfg, hoy, hace7):
    os.makedirs(OUT_DIR, exist_ok=True)
    ts    = hoy.strftime("%Y%m%d")
    fname = f"SIEG_OSINT_Informe_Semanal_{ts}.xlsx"
    fpath = os.path.join(OUT_DIR, fname)

    wb = Workbook()

    map_sent = {"POS": 1, "NEU": 0, "NEG": -1}
    df_noticias["sent_score"] = df_noticias["sentiment_label"].map(map_sent)

    # ── Hoja 1: Resumen ejecutivo ─────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_view.showGridLines = False

    ws1["A1"] = "SIEG – Centro OSINT · Política Nacional"
    ws1["A1"].font = Font(bold=True, size=16, color=COLOR_GREEN, name="Courier New")
    ws1["A2"] = f"Informe Semanal · {hace7.strftime('%d/%m/%Y')} → {hoy.strftime('%d/%m/%Y')}"
    ws1["A2"].font = Font(size=11, color=COLOR_GREEN, name="Courier New")
    ws1["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · © 2026 M. Castillo"
    ws1["A3"].font = Font(size=9, color="888888", italic=True)

    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A1:F1")
    ws1.merge_cells("A2:F2")
    ws1.merge_cells("A3:F3")

    # KPIs
    ws1["A5"] = "INDICADOR"
    ws1["B5"] = "VALOR"
    estilo_header(ws1, 5, 2)

    kpis = [
        ("Total noticias analizadas", len(df_noticias)),
        ("Partidos monitorizados", df_noticias["partido"].nunique()),
        ("Fuentes activas", df_noticias["source"].nunique()),
        ("Sentimiento medio (-1 a +1)", round(df_noticias["sent_score"].mean(), 3)),
        ("% Noticias negativas", f"{(df_noticias['sentiment_label']=='NEG').mean()*100:.1f}%"),
        ("% Noticias positivas", f"{(df_noticias['sentiment_label']=='POS').mean()*100:.1f}%"),
        ("Días cubiertos", df_noticias["fecha"].nunique()),
    ]

    for i, (label, val) in enumerate(kpis, start=6):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = val
        ws1[f"A{i}"].font = Font(name="Courier New", size=10)
        ws1[f"B{i}"].font = Font(bold=True, name="Courier New", size=10)

    # Top narrativas
    ws1["A14"] = "TOP NARRATIVAS DE LA SEMANA"
    ws1["A14"].font = Font(bold=True, color=COLOR_GREEN, name="Courier New")
    estilo_header(ws1, 14, 2)

    narr_counts = {}
    for _, row in df_noticias.iterrows():
        if row.get("narrativas") and str(row["narrativas"]) not in ["ninguna","nan"]:
            for n in str(row["narrativas"]).split(","):
                n = n.strip()
                if n:
                    narr_counts[n] = narr_counts.get(n, 0) + 1

    for i, (narr, cnt) in enumerate(sorted(narr_counts.items(), key=lambda x:-x[1])[:8], start=15):
        ws1[f"A{i}"] = narr
        ws1[f"B{i}"] = cnt
        ws1[f"A{i}"].font = Font(name="Courier New", size=10)
        ws1[f"B{i}"].font = Font(bold=True, name="Courier New", size=10)

    autofit(ws1)

    # ── Hoja 2: Noticias ──────────────────────────────────
    ws2 = wb.create_sheet("Noticias")
    ws2.sheet_view.showGridLines = False

    cols_noticias = ["fecha", "source", "source_type", "partido",
                     "sentiment_label", "sent_score", "narrativas", "temas", "title", "link"]
    cols_noticias = [c for c in cols_noticias if c in df_noticias.columns]

    for col_idx, col_name in enumerate(cols_noticias, 1):
        ws2.cell(1, col_idx, col_name.upper())
    estilo_header(ws2, 1, len(cols_noticias))

    for row_idx, row in df_noticias[cols_noticias].iterrows():
        for col_idx, val in enumerate(row, 1):
            cell = ws2.cell(row_idx + 2, col_idx, str(val) if val is not None else "")
            cell.font = Font(name="Courier New", size=9)
            # Color por sentimiento
            if col_idx == cols_noticias.index("sentiment_label") + 1 if "sentiment_label" in cols_noticias else -1:
                if str(val) == "POS":
                    cell.fill = PatternFill("solid", fgColor="E8F5E9")
                elif str(val) == "NEG":
                    cell.fill = PatternFill("solid", fgColor="FFEBEE")

    autofit(ws2)

    # ── Hoja 3: Valoración líderes ────────────────────────
    ws3 = wb.create_sheet("Valoracion Lideres")
    ws3.sheet_view.showGridLines = False

    lideres = cfg.get("lideres", {})
    headers3 = ["PARTIDO", "LÍDER", "POSICIÓN", "IDEOLOGÍA",
                 "NOTICIAS", "SENT. MEDIO", "% POS", "% NEG", "% NEU"]
    for col_idx, h in enumerate(headers3, 1):
        ws3.cell(1, col_idx, h)
    estilo_header(ws3, 1, len(headers3))

    ranking = []
    for partido, grupo in df_noticias.groupby("partido"):
        info = lideres.get(partido, {})
        sent = grupo["sent_score"].mean()
        ranking.append({
            "partido":  partido,
            "lider":    info.get("lider", "—"),
            "posicion": info.get("posicion", "—"),
            "ideologia":info.get("ideologia", "—"),
            "noticias": len(grupo),
            "sent":     round(sent, 3),
            "pct_pos":  round((grupo["sentiment_label"]=="POS").mean()*100, 1),
            "pct_neg":  round((grupo["sentiment_label"]=="NEG").mean()*100, 1),
            "pct_neu":  round((grupo["sentiment_label"]=="NEU").mean()*100, 1),
        })
    ranking.sort(key=lambda x: -x["sent"])

    for row_idx, r in enumerate(ranking, 2):
        vals = [r["partido"], r["lider"], r["posicion"], r["ideologia"],
                r["noticias"], r["sent"], r["pct_pos"], r["pct_neg"], r["pct_neu"]]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row_idx, col_idx, val)
            cell.font = Font(name="Courier New", size=10)
        # Color sentimiento
        sent_cell = ws3.cell(row_idx, 6)
        if r["sent"] > 0.1:
            sent_cell.fill = PatternFill("solid", fgColor="E8F5E9")
        elif r["sent"] < -0.1:
            sent_cell.fill = PatternFill("solid", fgColor="FFEBEE")

    autofit(ws3)

    # ── Hoja 4: Tendencias ────────────────────────────────
    ws4 = wb.create_sheet("Tendencias")
    ws4.sheet_view.showGridLines = False

    if not df_tendencias.empty:
        headers4 = [c.upper() for c in df_tendencias.columns]
        for col_idx, h in enumerate(headers4, 1):
            ws4.cell(1, col_idx, h)
        estilo_header(ws4, 1, len(headers4))

        for row_idx, row in df_tendencias.iterrows():
            for col_idx, val in enumerate(row, 1):
                cell = ws4.cell(row_idx + 2, col_idx, val)
                cell.font = Font(name="Courier New", size=9)

    autofit(ws4)

    # ── Hoja 5: Narrativas ────────────────────────────────
    ws5 = wb.create_sheet("Narrativas")
    ws5.sheet_view.showGridLines = False

    ws5.cell(1, 1, "NARRATIVA")
    ws5.cell(1, 2, "MENCIONES")
    ws5.cell(1, 3, "DESCRIPCIÓN")
    estilo_header(ws5, 1, 3)

    narr_perfil = cfg.get("narrativas_perfil", {})
    for row_idx, (narr, cnt) in enumerate(
        sorted(narr_counts.items(), key=lambda x: -x[1]), 2
    ):
        ws5.cell(row_idx, 1, narr).font = Font(name="Courier New", size=10, bold=True)
        ws5.cell(row_idx, 2, cnt).font = Font(name="Courier New", size=10)
        desc = narr_perfil.get(narr, {}).get("desc", "—")
        ws5.cell(row_idx, 3, desc).font = Font(name="Courier New", size=9, italic=True)

    autofit(ws5)

    wb.save(fpath)
    log(f"Excel generado: {fpath} ({os.path.getsize(fpath)//1024}KB)")
    return fpath, fname

def enviar_telegram(fpath, fname, hoy, hace7, df_noticias):
    """Envía el Excel por Telegram."""
    api = f"https://api.telegram.org/bot{BOT_TOKEN}"

    caption = (
        f"📊 <b>SIEG – Informe Semanal</b>\n"
        f"📅 {hace7.strftime('%d/%m/%Y')} → {hoy.strftime('%d/%m/%Y')}\n\n"
        f"📰 {len(df_noticias):,} noticias analizadas\n"
        f"🏛️ {df_noticias['partido'].nunique()} partidos\n"
        f"📡 {df_noticias['source'].nunique()} fuentes\n\n"
        f"📥 Descarga el Excel con datos completos\n"
        f"© 2026 M. Castillo · mybloggingnotes@gmail.com"
    )

    with open(fpath, "rb") as f:
        r = requests.post(f"{api}/sendDocument", data={
            "chat_id": CHANNEL,
            "caption": caption,
            "parse_mode": "HTML"
        }, files={"document": (fname, f)}, timeout=60)

    if r.status_code == 200:
        log("Excel enviado por Telegram OK")
    else:
        log(f"Error Telegram: {r.status_code}")

def main():
    log("=" * 50)
    log("Iniciando exportación semanal")

    hoy    = datetime.now().date()
    hace7  = hoy - timedelta(days=7)

    conn = sqlite3.connect(DB_PATH)

    df_noticias = pd.read_sql_query(
        "SELECT * FROM noticias_norm WHERE DATE(created_at) >= ?",
        conn, params=(str(hace7),)
    )
    df_noticias["fecha"] = pd.to_datetime(
        df_noticias["created_at"], errors="coerce"
    ).dt.date

    df_tendencias = pd.read_sql_query(
        "SELECT * FROM tendencias_diarias WHERE fecha >= ?",
        conn, params=(str(hace7),)
    )
    conn.close()

    with open(CFG_PATH, encoding="utf-8") as f:
        cfg = json.load(f)

    log(f"Datos: {len(df_noticias)} noticias, {len(df_tendencias)} tendencias")

    fpath, fname = crear_excel(df_noticias, df_tendencias, cfg, hoy, hace7)
    enviar_telegram(fpath, fname, hoy, hace7, df_noticias)

    # Copiar al export para Streamlit Cloud
    export_path = os.path.join(BASE_DIR, "data", "export", "ultimo_informe_semanal.xlsx")
    import shutil
    shutil.copy2(fpath, export_path)
    log(f"Copiado a: {export_path}")

    log("Exportación semanal completada")
    log("=" * 50)

if __name__ == "__main__":
    main()
